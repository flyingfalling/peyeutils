import openpyxl
import sys
import pandas as pd;
from openpyxl.styles import PatternFill

xlsx_file=sys.argv[1];
wajd_idx=sys.argv[2];


wb = openpyxl.load_workbook(xlsx_file);
ws = wb.active


# Define the fill pattern to check for (e.g., solid red)
#red_fill = PatternFill(patternType='solid', fgColor="FFFF0000") 

rowlist=list();
START_ROW=2; #first row header, and 1-indexed
# Iterate through all rows
highlightedrows=0;
colors=set();
for row in ws.iter_rows(min_row=START_ROW, max_row=ws.max_row):
    # Check if a specific cell in the row is highlighted
    if row[0].fill.start_color.index != "00000000":
        if( row[0].fill.start_color.index != 'FFFFFF00' ):
            print(f"WTF Row {row[0].row} is highlighted ZERO?!?!?!");
        rowlist.append(True);
        highlightedrows+=1;
        colors.add(row[0].fill.start_color.index);
        #print(f"Row {row[0].row} is highlighted.")
        pass;
    else:
        rowlist.append(False);
        #print(f"Row {row[0].row} is not highlighted.")
        pass;
    pass;

#print(colors);
#print(highlightedrows);
#exit(0);

df = pd.read_excel(xlsx_file);
cols=df.columns[0];
df[ cols.split(' ') ] = df[ cols ].str.split(' ', expand=True);
df = df.drop(columns=cols);

df['muscimol']=rowlist; #REV; first row was header...but skipped
print(df[df.muscimol==True]);
print(df);
df.to_csv('wajd_muscimol_labelled.csv', index=False);

wdf = pd.read_csv(wajd_idx);

nomusc_fn=wajd_idx + '.muscimol.csv';

wdf = pd.merge(left=wdf, left_on=['date','subj','trialcsv'],
               right=df, right_on=['date','subj','trialcsv'],
               how='outer' );

print(wdf);
print(wdf.columns);

wdf.to_csv(nomusc_fn, index=False);

